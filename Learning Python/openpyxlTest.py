import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('/Users/tony/Tony/example.xlsx')
print(wb.sheetnames)
sheet = wb['名单']
print(sheet.title)
print(sheet['A1'].value)
anotherSheet = wb.active
print(anotherSheet)
c = sheet['B2']
print(c.value)
print(c.parent)
tmp = 'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value
print(tmp)
tmp1 = 'Cell ' + c.coordinate + ' is ' + c.value
print(tmp1)
sheet.cell(row=2, column=2)
print(sheet.cell(row=2, column=3).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=2, column=i).value)
print(sheet.max_row)
print(sheet.max_column)
print(get_column_letter(1))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))

print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObject in rowOfCellObjects:
        print(cellObject.coordinate, cellObject.value)
    print('--- END OF ROW ---')

activeSheet = wb.active
print(activeSheet)
print(list(activeSheet.columns)[0])
